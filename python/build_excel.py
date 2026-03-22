"""
build_excel.py
Reads a JSON file of rows and builds a styled .xlsx with openpyxl.

Usage:
  python3 build_excel.py <input_json> <output_xlsx>
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build(input_json: str, output_xlsx: str):
    with open(input_json, encoding='utf-8') as f:
        data = json.load(f)

    rows    = data['rows']
    version = data.get('version', '')

    if not rows:
        raise ValueError("Tidak ada data rows")

    headers = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = version[:31]  # Excel sheet name max 31 chars

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    header_fill  = PatternFill('solid', start_color='1F4E79')  # dark blue
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font    = Font(name='Arial', size=10)
    data_align   = Alignment(horizontal='left', vertical='center', wrap_text=False)
    center_align = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),   right=Side(style='thin'),
        top=Side(style='thin'),    bottom=Side(style='thin'),
    )

    # Number columns (right-align)
    number_cols = {
        'FF Saham <5%', 'FF Direksi/Komisaris <5%', 'FF Pengendali <5%',
        'FF Afiliasi <5%', 'FF Treasury <5%', 'FF Porto Investasi',
        'FreeFloat Total', 'FF Saham Tercatat', '%FF',
        'Total Pengendali Bulan Ini', 'Total Non Pengendali Bulan Ini',
        'Persentase Total Pengendali Bulan Ini',
        'Persentase Total Non Pengendali Bulan Ini',
        'Pemegang Saham',
    }

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 18
        for ci, h in enumerate(headers, 1):
            val  = row.get(h, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font   = data_font
            cell.border = thin_border

            if h in number_cols:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                # Format number cells
                if isinstance(val, (int, float)) and val != '':
                    if h == '%FF' or 'Persentase' in h:
                        cell.number_format = '0.000'
                    else:
                        cell.number_format = '#,##0'
            else:
                cell.alignment = data_align

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        'Month': 12,
        'Date of PDF': 18,
        'Kode Emiten': 12,
        'Nomor Surat': 28,
        'FF Saham <5%': 16,
        'FF Direksi/Komisaris <5%': 22,
        'FF Pengendali <5%': 18,
        'FF Afiliasi <5%': 16,
        'FF Treasury <5%': 16,
        'FF Porto Investasi': 18,
        'FreeFloat Total': 16,
        'FF Saham Tercatat': 18,
        '%FF': 8,
        'Total Pengendali Bulan Ini': 24,
        'Total Non Pengendali Bulan Ini': 28,
        'Persentase Total Pengendali Bulan Ini': 34,
        'Persentase Total Non Pengendali Bulan Ini': 38,
        'Pemegang Saham': 16,
        'Penerima Manfaat Akhir': 50,
        'Penerima Manfaat Akhir (PT)': 35,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 16)

    # ── Freeze top row ────────────────────────────────────────────────────────
    ws.freeze_panes = 'A2'

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    wb.save(output_xlsx)
    print(f'OK: {output_xlsx} ({len(rows)} rows)')

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: build_excel.py <input_json> <output_xlsx>')
        sys.exit(1)
    build(sys.argv[1], sys.argv[2])